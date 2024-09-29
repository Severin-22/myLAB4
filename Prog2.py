import pandas as pd  
from openpyxl import Workbook  
from openpyxl.utils.dataframe import dataframe_to_rows  
from openpyxl.styles import Font, Alignment  
from datetime import datetime 
import os  

def calculate_age(birth_date_str):  #  функція обчислює вік на основі дати народження
    # Спробую перетворити рядок дати в формат datetime
    try:
        birth_date = datetime.fromisoformat(birth_date_str)
    except ValueError:  
        return None  

    today = datetime.today()  # Отримую сьогоднішню дату
    # вираховую вік
    age = today.year - birth_date.year - ((today.month, today.day) < (birth_date.month, birth_date.day))  
    return age  

def categorize_records(df):
    # очищення заголовків стовпців від зайвих пробілів
    df.columns = df.columns.str.strip()

    # перевірка чи є стовпець 'Дата народження' в даних
    if 'Дата народження' not in df.columns:
        print("Стовпець 'Дата народження' не знайдено.")
        return None  

    # додаю новий стовпець 'Вік', де обчислюю вік для кожного рядка
    df['Вік'] = df['Дата народження'].apply(calculate_age)

    # розподілення за віком
    younger_18 = df[df['Вік'] < 18]
    age_18_45 = df[(df['Вік'] >= 18) & (df['Вік'] <= 45)]
    age_45_70 = df[(df['Вік'] > 45) & (df['Вік'] <= 70)]
    older_70 = df[df['Вік'] > 70]

    # Повертаю словник з усіма категоріями
    return {
        "all": df,
        "younger_18": younger_18,
        "18-45": age_18_45,
        "45-70": age_45_70,
        "older_70": older_70
    }

def write_xlsx(filename, sheets):
    # Функція для запису даних у Excel файл
    try:
        wb = Workbook()  # Створюю новий Excel файл

        for sheet_name, data in sheets.items():
            # Для кожного аркуша в словнику створюю новий аркуш
            ws = wb.create_sheet(title=sheet_name)
            # Додаю дані в аркуш
            for r in dataframe_to_rows(data, index=False, header=True):
                ws.append(r)

            # Форматую заголовки, щоб вони виглядали краще
            for cell in ws[1]:  # Перший рядок - заголовки
                cell.font = Font(bold=True)  # Роблю текст заголовків жирним
                cell.alignment = Alignment(horizontal="center")  # Вирівнюю текст по центру

        # Якщо стандартний аркуш існує, я його видаляю
        if 'Sheet' in wb.sheetnames:
            del wb['Sheet']

        # Зберігаю Excel файл
        wb.save(filename)
        print("Дані успішно відформатовані!")  # Повідомлення про успіх
        print("Ok")

    except Exception as e:
        # Якщо сталася помилка, я виводжу повідомлення про неї
        print(f"Помилка створення XLSX файлу: {e}")

def main():
    csv_file = 'Serhii-table.csv'  # Ім'я CSV файлу
    xlsx_file = 'employees.xlsx'  # Ім'я Excel файлу для збереження

    # Перевіряю, чи існує CSV файл
    if not os.path.isfile(csv_file):
        print("Помилка відкриття файлу CSV.")
        return

    try:
        # Читаю CSV файл
        df = pd.read_csv(csv_file, sep=',', encoding='utf-8-sig')  # Змінюю sep на ',' для розділення
        print("Заголовки стовпців:", df.columns)  # Виводжу заголовки стовпців
        print("Перші 5 рядків даних:\n", df.head())  # Виводжу перші 5 рядків даних
    except Exception as e:
        # Якщо сталася помилка, я виводжу повідомлення про неї
        print(f"Повідомлення про відсутність або проблеми при відкритті файлу CSV: {e}")
        return

    # Перевіряю, чи існує стовпець 'Дата народження' після завантаження
    if 'Дата народження' not in df.columns:
        print("Стовпець 'Дата народження' не знайдено. Будь ласка, перевірте заголовки.")
        return

    sheets = categorize_records(df)  
    write_xlsx(xlsx_file, sheets)  #записую дані у Excel файл

if __name__ == "__main__":
    main() 
